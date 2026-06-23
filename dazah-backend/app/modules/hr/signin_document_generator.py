"""Generate training sign-in sheet documents from templates."""

from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font

from app.modules.hr.schemas import TrainingSignInSheetInput


def _fmt_date(value) -> str:
    """将日期格式化为 YYYY.MM.DD."""
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y.%m.%d")
    if isinstance(value, str):
        from datetime import datetime
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(value, fmt).strftime("%Y.%m.%d")
            except ValueError:
                continue
    return str(value)


OLD_TEMPLATE_NAME = "7.5培训签到表.xlsx"
NEW_TEMPLATE_NAME = "R-GN-2002 K 培训签到表.xlsx"


def _find_old_template() -> Path:
    """Locate the old factory xlsx template."""
    candidates = [
        Path("员工培训教育管理规程") / OLD_TEMPLATE_NAME,
        Path("../员工培训教育管理规程") / OLD_TEMPLATE_NAME,
        Path(__file__).resolve().parent.parent.parent.parent
        / "员工培训教育管理规程"
        / OLD_TEMPLATE_NAME,
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(f"模板文件未找到: {OLD_TEMPLATE_NAME}")


def _find_new_template() -> Path:
    """Locate the new factory xls template."""
    candidates = [
        Path("新厂人员培训管理规程") / NEW_TEMPLATE_NAME,
        Path("../新厂人员培训管理规程") / NEW_TEMPLATE_NAME,
        Path(__file__).resolve().parent.parent.parent.parent
        / "新厂人员培训管理规程"
        / NEW_TEMPLATE_NAME,
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(f"模板文件未找到: {NEW_TEMPLATE_NAME}")


def _generate_old(data: TrainingSignInSheetInput, page: int = 0) -> BytesIO:
    """Fill the old factory training sign-in sheet xlsx template."""
    template_path = _find_old_template()
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active

    # Training date: D4=year, F4=month, H4=day
    if data.training_date:
        date_parts = str(data.training_date).split("-")
        if len(date_parts) == 3:
            ws["D4"] = date_parts[0]  # year
            ws["F4"] = date_parts[1]  # month
            ws["H4"] = date_parts[2]  # day

    # Department: D5:I5 merged
    if data.department:
        ws["D5"] = data.department

    # Training method: tick the checkbox in J5:N5
    if data.training_method:
        method_map = {
            "面授": "□面授",
            "函授": "□函授",
            "远程教育": "□远程教育",
            "自学": "□自学",
            "其他": "□其他：",
        }
        j5_value = ws["J5"].value or ""
        for label, placeholder in method_map.items():
            if label in data.training_method:
                j5_value = j5_value.replace(placeholder, f"☑{placeholder[1:]}")
        ws["J5"] = j5_value

    # Attendance count: J6:N6 merged
    total = len(data.employee_names)
    ws["J6"] = f"应受训人数：{total}人   实际受训人数合计：      人"

    # Training time: A8=start_hour, C8=start_min, E8=end_hour, G8=end_min
    if data.training_time_start and data.training_time_end:
        start_parts = data.training_time_start.split(":")
        end_parts = data.training_time_end.split(":")
        if len(start_parts) == 2:
            ws["A8"] = start_parts[0]
            ws["C8"] = start_parts[1]
        if len(end_parts) == 2:
            ws["E8"] = end_parts[0]
            ws["G8"] = end_parts[1]

    # Topic: H8:M8 merged (below the header H7:M7)
    display_topic = data.topic
    if data.training_subject:
        display_topic = f"{data.training_subject} — {data.topic}"
    ws["H8"] = display_topic
    ws["H8"].alignment = Alignment(horizontal="center", vertical="center")

    # Instructor: N8:N12 merged (below the header N7)
    if data.instructor:
        ws["N8"] = data.instructor
        ws["N8"].alignment = Alignment(horizontal="center", vertical="center")

    # Remarks: A30:N30 merged
    if data.remarks:
        ws["A30"] = f"备注：{data.remarks}"

    # Employee names list (max 30 per page)
    page_size = 30
    start = page * page_size
    page_names = data.employee_names[start : start + page_size]

    # Clear existing numeric placeholders first
    for row in range(15, 30):
        for col in ["A", "K"]:
            cell = ws[f"{col}{row}"]
            if cell and cell.value and str(cell.value).strip().isdigit():
                cell.value = ""

    for i, name in enumerate(page_names):
        if i < 15:
            row = 15 + i
            cell = ws[f"A{row}"]
            if cell:
                cell.value = name
                cell.font = Font(name="宋体", size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            row = 15 + (i - 15)
            cell = ws[f"K{row}"]
            if cell:
                cell.value = name
                cell.font = Font(name="宋体", size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _generate_new(data: TrainingSignInSheetInput, page: int = 0) -> BytesIO:
    """Fill the new factory training sign-in sheet xlsx template."""
    template_path = _find_new_template()
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active

    # Row 5 (A5:E5 merged): 培训日期
    if data.training_date:
        ws["A5"] = f"培训日期：{_fmt_date(data.training_date)}"

    # Row 6 (A6:E6 merged): 培训方式 — 勾选对应项
    if data.training_method:
        method_map = {
            "面授": "培训方式：☑ 面授 □ 函授 □ 远程教育 □ 自学 □其他：",
            "函授": "培训方式：□ 面授 ☑ 函授 □ 远程教育 □ 自学 □其他：",
            "远程教育": "培训方式：□ 面授 □ 函授 ☑ 远程教育 □ 自学 □其他：",
            "自学": "培训方式：□ 面授 □ 函授 □ 远程教育 ☑ 自学 □其他：",
            "其他": "培训方式：□ 面授 □ 函授 □ 远程教育 □ 自学 ☑其他：",
        }
        ws["A6"] = method_map.get(data.training_method, f"培训方式：{data.training_method}")

    # Row 7 (A7:E7 merged): 受训部门/班组
    if data.department:
        ws["A7"] = f"受训部门/班组：{data.department}"

    # Row 8 (A8:E8 merged): 应受训人数 / 实际受训人数
    total = len(data.employee_names)
    ws["A8"] = f"应受训人数：{total} 人           实际受训人数合计：      人"

    # Row 10: 培训时间 / 培训题目 / 授课人
    # A10, A11, A12 — use first row for data
    if data.training_time_start and data.training_time_end:
        ws["A10"] = f"{data.training_time_start} ~ {data.training_time_end}"
    if data.topic:
        ws["B10"] = data.topic
    if data.instructor:
        ws["E10"] = data.instructor

    # Clear placeholder rows A11, A12
    ws["A11"] = ""
    ws["A12"] = ""

    # Employee names (max 30 per page, starting from row 15)
    page_size = 30
    start = page * page_size
    page_names = data.employee_names[start : start + page_size]

    for i, name in enumerate(page_names):
        if i < 15:
            ws.cell(row=15 + i, column=1).value = name  # A15-A29
        else:
            ws.cell(row=15 + (i - 15), column=3).value = name  # C15-C29

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_training_sign_in_sheet(data: TrainingSignInSheetInput, factory: str = "old", page: int = 0) -> BytesIO:
    """Fill the training sign-in sheet template with form data.

    Each page holds up to 30 employees. Returns a BytesIO buffer
    containing the generated document for the requested page.
    """
    if factory == "new":
        return _generate_new(data, page)
    return _generate_old(data, page)
